import json
import sys
from pathlib import Path
import openpyxl

# Claude Code（WSL環境）での文字化け防止
sys.stdout.reconfigure(encoding="utf-8")

# =============================================
# .envファイルの読み込み（ルートから実行想定）
# =============================================

def load_env(env_path: Path) -> dict:
    if not env_path.exists():
        raise FileNotFoundError(f".envファイルが見つかりません: {env_path}")
    env = {}
    for line in env_path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.strip().startswith("#"):
            continue
        key, _, val = line.partition("=")
        env[key.strip()] = val.strip()
    return env

env = load_env(Path(".env"))

# =============================================
# Excel (MemberList テーブル) → JSON 変換
# ヘッダー: 2行目 / ボディ: 3行目以降
# =============================================

BASE_DIR    = Path(env["BASE_DIR"])
EXCEL_PATH  = BASE_DIR / env["EXCEL_FILE"]
OUTPUT_PATH = BASE_DIR / env["OUTPUT_FILE"]
SHEET_NAME  = env["SHEET_NAME"]
HEADER_ROW     = 2
DATA_START_ROW = 3

# --- Excel 読み込み ---
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb[SHEET_NAME]

# ヘッダー取得（2行目・空セルで終了）
headers = []
for cell in ws[HEADER_ROW]:
    if cell.value is None:
        break
    headers.append(str(cell.value).strip())

print(f"検出されたヘッダー ({len(headers)} 列): {', '.join(headers)}")

# ボディ取得（3行目以降・全列空で終了）
records = []
for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=len(headers), values_only=True):
    if all(v is None for v in row):
        break
    record = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
    records.append(record)

print(f"{len(records)} 件のレコードを取得しました。")

# JSON 出力
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

print(f"JSON 出力完了: {OUTPUT_PATH}")